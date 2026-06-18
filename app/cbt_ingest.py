# -*- coding: utf-8 -*-
"""CBT-FULL 官方导出(B) → 飞书报表 逐SKU全口径解析 (云端版, 2026-06-18 task3云化).

背景: CBT API 被 429 墙在 ~85% 缓存 + Full仓储无 API → 纯 API 做不到准确 CBT 毛利.
官方导出(Orders+账单+广告)是唯一完整准确源(B). 俊辉每月把3文件传飞书云盘文件夹 → 本模块下载解析 →
按SKU update 飞书 CBT 行(保留美通头程/海外仓列). 本地版前身 C:/tmp/cbt_export_ingest.py.

口径(USD): 锚定导出 S列(净受领, 已含运费收入N/汇率M/额外运费P). 全额毛利=飞书公式字段(本模块不写).
RMB = USD × fx. 采购 = 领星 cg_price × units(走 lingxing).
"""
import io, os, json, time, re
from collections import defaultdict
import httpx
import openpyxl

from app import lingxing

FEISHU = "https://open.feishu.cn/open-apis"
APP_TOKEN = "WM3LbBr76aRqMys2of8c1dGInEb"
TABLE_ID = "tbl09sRPkX35PDfU"
SHOP_LABEL = "ML CBT-FULL (1502236229)"
F_FULL = "Full仓储费(RMB)"

# Orders report 列(0-based, row8起数据). N(13)=运费收入(S列已含,不单列), S(18)=净受领锚点.
O_UNITS, O_K, O_L, O_N, O_O, O_Q, O_R, O_S, O_SKU, O_LISTING, O_TITLE = 9,10,11,13,14,16,17,18,24,25,26


async def _ft() -> str:
    app_id = os.getenv("FEISHU_APP_ID", "cli_a9f6ae86fce8dbd8")
    secret = os.getenv("FEISHU_APP_SECRET", "")
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(f"{FEISHU}/auth/v3/tenant_access_token/internal",
                         json={"app_id": app_id, "app_secret": secret})
    return r.json()["tenant_access_token"]


async def _list_folder(tok: str, folder_token: str) -> list[dict]:
    out = []
    page = None
    async with httpx.AsyncClient(timeout=30) as c:
        while True:
            url = f"{FEISHU}/drive/v1/files?folder_token={folder_token}&page_size=50" + (f"&page_token={page}" if page else "")
            d = (await c.get(url, headers={"Authorization": f"Bearer {tok}"})).json().get("data", {})
            out += d.get("files", [])
            page = d.get("next_page_token")
            if not d.get("has_more") or not page:
                break
    return out


async def _download(tok: str, file_token: str) -> bytes:
    async with httpx.AsyncClient(timeout=120) as c:
        r = await c.get(f"{FEISHU}/drive/v1/files/{file_token}/download", headers={"Authorization": f"Bearer {tok}"})
        r.raise_for_status()
        return r.content


def _pick(files: list[dict], *kws) -> dict | None:
    # 取名字含关键词且 type=file 的最新一个(按 modified_time 倒序; 容多次上传取最新)
    cands = [f for f in files if f.get("type") == "file" and any(k in (f.get("name") or "") for k in kws)]
    cands.sort(key=lambda f: f.get("modified_time", "0"), reverse=True)
    return cands[0] if cands else None


def _num(r, i):
    return r[i] if (i < len(r) and isinstance(r[i], (int, float))) else 0.0


def _parse_orders(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Orders US"] if "Orders US" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True))
    wb.close()
    listing2sku = {}
    for r in rows:
        if r[O_SKU] and r[O_LISTING]:
            listing2sku[str(r[O_LISTING])] = str(r[O_SKU])
    agg = defaultdict(lambda: dict(units=0.0, K=0.0, L=0.0, N=0.0, O=0.0, Q=0.0, R=0.0, S=0.0, orders=0, title=""))
    for r in rows:
        sku = r[O_SKU]
        if not sku:
            sku = listing2sku.get(str(r[O_LISTING])) or "(组合包)"
        a = agg[str(sku)]
        a["units"] += _num(r, O_UNITS); a["K"] += _num(r, O_K); a["L"] += _num(r, O_L)
        a["N"] += _num(r, O_N); a["O"] += _num(r, O_O); a["Q"] += _num(r, O_Q)
        a["R"] += _num(r, O_R); a["S"] += _num(r, O_S); a["orders"] += 1
        if not a["title"] and r[O_TITLE]:
            a["title"] = str(r[O_TITLE])
    return agg, listing2sku


def _parse_ads(data: bytes, listing2sku: dict):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Report by ads"] if "Report by ads" in wb.sheetnames else wb.worksheets[-1]
    ad = defaultdict(float); total = 0.0; unmatched = 0.0
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        lst = r[4] if len(r) > 4 else None
        inv = r[12] if (len(r) > 12 and isinstance(r[12], (int, float))) else 0.0
        if not inv:
            continue
        total += inv
        sku = listing2sku.get(str(lst)) if lst else None
        if sku:
            ad[sku] += inv
        else:
            unmatched += inv
    wb.close()
    return dict(ad), total, unmatched


def _parse_storage(data: bytes) -> float:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["REPORT"]
    KW = ["仓陈旧", "Long-term storage", "存储服务费", "入库货件违约", "Storage service fee", "almacenamiento"]
    total = 0.0
    for r in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
        d = r[3] if len(r) > 3 else None
        v = r[7] if (len(r) > 7 and isinstance(r[7], (int, float))) else 0.0
        if d is None or not v:
            continue
        if any(k in str(d) for k in KW):
            total += v
    wb.close()
    return total


async def _bitable_all(tok: str) -> list[dict]:
    out = []
    body = {"page_size": 500}
    async with httpx.AsyncClient(timeout=60) as c:
        while True:
            d = (await c.post(f"{FEISHU}/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records/search",
                              json=body, headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"})).json().get("data", {})
            out += d.get("items", [])
            if d.get("has_more") and d.get("page_token"):
                body["page_token"] = d["page_token"]
            else:
                break
    return out


async def _ensure_full_field(tok: str):
    async with httpx.AsyncClient(timeout=30) as c:
        d = (await c.get(f"{FEISHU}/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/fields?page_size=200",
                         headers={"Authorization": f"Bearer {tok}"})).json().get("data", {})
        if F_FULL not in {f["field_name"] for f in d.get("items", [])}:
            await c.post(f"{FEISHU}/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/fields",
                         json={"field_name": F_FULL, "type": 2},
                         headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"})


async def run(month: str, commit: bool = False, fx: float = 6.8628,
              folder_token: str | None = None) -> dict:
    folder_token = folder_token or os.getenv("CBT_EXPORT_FOLDER_TOKEN", "")
    if not folder_token:
        return {"status": "error", "msg": "CBT_EXPORT_FOLDER_TOKEN 未配置"}
    period = f"month_{month}"
    tok = await _ft()
    files = await _list_folder(tok, folder_token)
    f_ord = _pick(files, "Orders", "Mercado_Libre")
    f_ad = _pick(files, "report-pads", "广告", "pads")
    f_bill = _pick(files, "BILL", "账单")
    if not f_ord:
        return {"status": "error", "msg": "飞书云盘文件夹找不到 Orders 导出(名字含 Orders/Mercado_Libre)",
                "folder_files": [f.get("name") for f in files]}

    agg, l2s = _parse_orders(await _download(tok, f_ord["token"]))
    ad_sku, ad_total, ad_unmatched = (_parse_ads(await _download(tok, f_ad["token"]), l2s) if f_ad else ({}, 0.0, 0.0))
    storage_total = (_parse_storage(await _download(tok, f_bill["token"])) if f_bill else 0.0)

    # 组合包空SKU行(0件)按营收摊到真实SKU
    blanks = [k for k in agg if str(k).strip() == "" or k == "(组合包)"]
    if blanks:
        real = [(k, v) for k, v in agg.items() if k not in blanks]
        rk = sum(v["K"] for _, v in real) or 1
        for bk in blanks:
            b = agg.pop(bk)
            for k, v in real:
                sh = v["K"] / rk
                for fld in ("K", "L", "N", "O", "Q", "R", "S"):
                    v[fld] += b[fld] * sh

    # 领星 cg
    products = await lingxing.fetch_all_products()
    def cg_for(sku):
        p = products.get(lingxing.resolve_erp_sku(sku))
        try:
            return float(p["cg_price"]) if p and p.get("cg_price") is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    tot_K = sum(a["K"] for a in agg.values()) or 1
    rows_out = []
    for sku, a in agg.items():
        K, L, Q, R, S = a["K"], -a["L"], -a["Q"], -a["R"], a["S"]
        O = a["K"] + a["L"] + a["Q"] + a["R"] - a["S"]  # net 物流(反推闭合到S)
        share = K / tot_K
        ad_alloc = ad_sku.get(sku, 0.0) + ad_unmatched * share
        storage_alloc = storage_total * share
        caigou = cg_for(sku) * a["units"]
        rows_out.append(dict(sku=sku, units=a["units"], orders=a["orders"], title=a["title"],
                             K=K, L=L, O=O, Q=Q, R=R, ad=ad_alloc, storage=storage_alloc, caigou=caigou,
                             has_cg=(cg_for(sku) > 0)))

    tot = {k: round(sum(r[k] for r in rows_out), 2) for k in ("K", "L", "O", "Q", "R", "ad", "storage", "caigou")}
    tot["units"] = int(sum(r["units"] for r in rows_out))
    summary = {"status": "ok", "month": month, "commit": commit, "fx": fx,
               "files": {"orders": f_ord["name"], "ads": (f_ad or {}).get("name"), "bill": (f_bill or {}).get("name")},
               "sku_count": len(rows_out), "totals": tot,
               "nocg_skus": [r["sku"] for r in rows_out if not r["has_cg"] and r["units"] > 0]}
    if not commit:
        summary["note"] = "DRY-RUN 未写飞书; commit=true 才按SKU update"
        return summary

    # 按SKU update 飞书(保留美通头程/海外仓; 全额毛利是公式字段不写)
    await _ensure_full_field(tok)
    existing = {}
    def _txt(v): return v[0].get("text", "") if isinstance(v, list) and v else (str(v) if v is not None else "")
    for it in await _bitable_all(tok):
        f = it["fields"]
        if "CBT" in _txt(f.get("店铺")) and _txt(f.get("周期")) == period:
            existing[_txt(f.get("SKU"))] = it["record_id"]
    now_ms = int(time.time() * 1000)
    upd = crt = 0; seen = set()
    async with httpx.AsyncClient(timeout=60) as c:
        H = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
        for r in rows_out:
            sku = r["sku"]; seen.add(sku)
            fld = {
                "SKU": sku, "平台": "Mercado Libre", "店铺": SHOP_LABEL, "周期": period,
                "订单数": int(r["orders"]), "件数": int(r["units"]), "币种": "USD", "我的汇率": round(fx, 4),
                "营收(原币)": round(r["K"], 2), "营收(RMB)": round(r["K"]*fx, 2),
                "客单价(原币)": round(r["K"]/r["orders"], 2) if r["orders"] else 0,
                "ML佣金(原币)": round(r["L"], 2), "ML佣金(RMB)": round(r["L"]*fx, 2),
                "物流费(原币)": round(r["O"], 2), "物流费(RMB)": round(r["O"]*fx, 2),
                "VAT估算(原币)": round(r["Q"], 2), "VAT估算(RMB)": round(r["Q"]*fx, 2),
                "退款金额(原币)": round(r["R"], 2), "退款金额(RMB)": round(r["R"]*fx, 2),
                "广告费(原币)": round(r["ad"], 2), "广告费(RMB)": round(r["ad"]*fx, 2),
                "采购成本(RMB)": round(r["caigou"], 2), F_FULL: round(r["storage"]*fx, 2),
                "卖家折扣(原币)": 0, "卖家折扣(RMB)": 0,
                "简易毛利(RMB)": round(r["K"]*fx - r["caigou"], 2),
                "商品标题": r["title"], "数据拉取时间": now_ms,
            }  # 🚨 不写 全额毛利(公式字段)/头程/海外仓(美通列保留)
            rid = existing.get(sku)
            if rid:
                await c.put(f"{FEISHU}/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records/{rid}", json={"fields": fld}, headers=H); upd += 1
            else:
                await c.post(f"{FEISHU}/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records", json={"fields": fld}, headers=H); crt += 1
    summary.update({"updated": upd, "created": crt,
                    "stale_in_feishu_not_export": [s for s in existing if s not in seen]})
    return summary
